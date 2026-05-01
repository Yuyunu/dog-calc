"""dog-calc v2 build script
從 狗狗鮮食計算表(新版工作版)_20260501_Final.xlsx 自動產生:
- foods.json (73 食材營養資料 per gram + metadata)
- standards.json (AAFCO/NRC 標準 + 比例分析定義)

執行方式: python3 build.py
"""
import json
import sys
import os
from openpyxl import load_workbook

XLSX = os.environ.get("XLSX_PATH", "../狗狗鮮食計算表(新版工作版)_20260501_Final.xlsx")
OUT_FOODS = "foods.json"
OUT_STANDARDS = "standards.json"

# Schema mapping: food DB col → JSON key (per gram)
SCHEMA = [
    ("D", "kcal", "熱量(kcal)"),
    ("E", "protein", "蛋白質(g)"),
    ("F", "fat", "脂肪(g)"),
    ("G", "carb", "碳水(g)"),
    ("H", "fiber", "膳食纖維(g)"),
    ("I", "ca_mg", "鈣(mg)"),
    ("J", "p_mg", "磷(mg)"),
    ("K", "k_mg", "鉀(mg)"),
    ("L", "na_mg", "鈉(mg)"),
    ("M", "mg_mg", "鎂(mg)"),
    ("N", "fe_mg", "鐵(mg)"),
    ("O", "zn_mg", "鋅(mg)"),
    ("P", "cu_mg", "銅(mg)"),
    ("Q", "mn_mg", "錳(mg)"),
    ("R", "iodine_ug", "碘(μg)"),
    ("S", "se_ug", "硒(μg)"),
    ("T", "vita_iu", "維生素A(IU)"),
    ("U", "vita_rae_ug", "維生素A(mcg RAE)"),
    ("V", "vitd_iu", "維生素D(IU)"),
    ("W", "vite_mg", "維生素E(mg)"),
    ("X", "vite_iu", "維生素E(IU)"),
    ("Y", "b1_mg", "維生素B1(mg)"),
    ("Z", "b2_mg", "維生素B2(mg)"),
    ("AA", "b3_mg", "維生素B3(mg)"),
    ("AB", "b5_mg", "維生素B5(mg)"),
    ("AC", "b6_mg", "維生素B6(mg)"),
    ("AD", "b9_ug", "維生素B9(μg)"),
    ("AE", "b12_ug", "維生素B12(μg)"),
    ("AF", "choline_mg", "膽鹼(mg)"),
    ("AG", "omega6_g", "Omega-6(g)"),
    ("AH", "omega3_g", "Omega-3(g)"),
    ("AI", "arg_g", "精胺酸(g)"),
    ("AJ", "his_g", "組胺酸(g)"),
    ("AK", "ile_g", "異白胺酸(g)"),
    ("AL", "leu_g", "白胺酸(g)"),
    ("AM", "lys_g", "離胺酸(g)"),
    ("AN", "met_g", "甲硫胺酸(g)"),
    ("AO", "cys_g", "半胱胺酸(g)"),
    ("AP", "phe_g", "苯丙胺酸(g)"),
    ("AQ", "tyr_g", "酪胺酸(g)"),
    ("AR", "thr_g", "蘇胺酸(g)"),
    ("AS", "trp_g", "色胺酸(g)"),
    ("AT", "val_g", "纈胺酸(g)"),
    ("AU", "water_g", "水分(g)"),
    ("AW", "taurine_g", "牛磺酸(g)"),
]

def col_to_idx(col):
    """A→1, AA→27"""
    n = 0
    for c in col:
        n = n * 26 + (ord(c) - ord("A") + 1)
    return n

def to_num(v):
    if v is None or v == "" or v == "—": return 0
    try: return float(v)
    except: return 0

def main():
    print(f"Reading: {XLSX}")
    if not os.path.exists(XLSX):
        print(f"ERROR: file not found: {XLSX}")
        sys.exit(1)
    wb = load_workbook(XLSX, data_only=True)
    
    # === foods.json ===
    ws = wb["食材營養資料庫"]
    foods = []
    for row in range(3, 74):
        name = ws.cell(row=row, column=1).value
        if not name: continue
        en = ws.cell(row=row, column=2).value or ""
        category = ws.cell(row=row, column=3).value or ""
        source = ws.cell(row=row, column=col_to_idx("AV")).value or ""
        notes = ws.cell(row=row, column=col_to_idx("AX")).value or ""
        grams_per_unit = to_num(ws.cell(row=row, column=col_to_idx("AY")).value) or 1
        
        # Determine display unit based on category
        if "(顆)" in category:
            unit = "顆"
        elif "(包)" in category:
            unit = "包"
        elif "(匙)" in category:
            unit = "匙"
        else:
            unit = "g"
        
        food = {
            "row": row,
            "name": name,
            "en": en,
            "category": category,
            "unit": unit,
            "gramsPerUnit": grams_per_unit,
            "source": source,
            "notes": notes,
        }
        for col, key, _label in SCHEMA:
            food[key] = to_num(ws.cell(row=row, column=col_to_idx(col)).value)
        foods.append(food)
    
    with open(OUT_FOODS, "w", encoding="utf-8") as f:
        json.dump(foods, f, ensure_ascii=False, indent=1)
    print(f"✓ {OUT_FOODS}: {len(foods)} ingredients")
    
    # === standards.json ===
    ws_dog = wb["狗狗資料與每日需求"]
    
    # Read AAFCO/NRC standards (rows 24-69)
    # Layout: A=name, B=unit, C=AAFCO_min, D=NRC, E=AAFCO_max, F=daily_min, G=daily_rec, H=daily_max
    standards_nutrients = []
    
    nutrient_key_map = {
        "蛋白質": "protein",
        "脂肪": "fat",
        "Omega-6 脂肪酸 (LA)": "omega6_g",
        "Omega-3 脂肪酸 (ALA)": "omega3_g",
        "EPA + DHA": "epa_dha",
        "鈣": "ca_mg",
        "磷": "p_mg",
        "鉀": "k_mg",
        "鈉": "na_mg",
        "氯": "cl_g",
        "鎂": "mg_mg",
        "鐵": "fe_mg",
        "銅": "cu_mg",
        "鋅": "zn_mg",
        "錳": "mn_mg",
        "碘": "iodine_ug",
        "硒": "se_ug",
        "維生素 A": "vita_iu",
        "維生素 D": "vitd_iu",
        "維生素 E": "vite_iu",
        "維生素 K": "vitk_mg",
        "維生素 B1 (硫胺素)": "b1_mg",
        "維生素 B2 (核黃素)": "b2_mg",
        "維生素 B3 (菸鹼酸)": "b3_mg",
        "維生素 B5 (泛酸)": "b5_mg",
        "維生素 B6 (吡哆醇)": "b6_mg",
        "維生素 B9 (葉酸)": "b9_ug",
        "維生素 B12": "b12_ug",
        "膽鹼": "choline_mg",
        "精胺酸": "arg_g",
        "組胺酸": "his_g",
        "異白胺酸": "ile_g",
        "白胺酸": "leu_g",
        "離胺酸": "lys_g",
        "甲硫胺酸": "met_g",
        "甲硫胺酸+半胱胺酸": "met_cys_g",
        "苯丙胺酸": "phe_g",
        "苯丙胺酸+酪胺酸": "phe_tyr_g",
        "蘇胺酸": "thr_g",
        "色胺酸": "trp_g",
        "纈胺酸": "val_g",
    }
    
    for r in range(23, 70):
        a = ws_dog.cell(row=r, column=1).value
        if not a or "▎" in str(a) or "【" in str(a) or a in ("營養素", "鈣磷比"): continue
        if a not in nutrient_key_map: continue
        
        b = ws_dog.cell(row=r, column=2).value or ""
        c = ws_dog.cell(row=r, column=3).value
        d = ws_dog.cell(row=r, column=4).value
        e = ws_dog.cell(row=r, column=5).value
        
        std = {
            "key": nutrient_key_map[a],
            "name": a,
            "unit": b,
            "aafco_min_per_1000kcal": to_num(c) if c not in (None, "", "—") else None,
            "nrc_per_1000kcal": to_num(d) if d not in (None, "", "—") else None,
            "aafco_max_per_1000kcal": to_num(e) if e not in (None, "", "—") else None,
        }
        standards_nutrients.append(std)
    
    # Activity coefficients (label 含 ×係數)
    activities = []
    for r in range(14, 20):
        s = ws_dog.cell(row=r, column=1).value
        v = ws_dog.cell(row=r, column=2).value
        d = ws_dog.cell(row=r, column=3).value
        if s and v:
            v_num = to_num(v)
            v_str = f"{v_num:g}"  # 1.6 / 2 / 3 等
            if d:
                label = f"{s} ×{v_str}（{d}）"
            else:
                label = f"{s} ×{v_str}"
            activities.append({"value": v_num, "label": label})
    
    standards = {
        "rer_formula": "70 * weight^0.75",
        "der_formula": "RER * activity_coef",
        "default_weight": 11.5,
        "default_activity": 1.2,
        "activity_options": activities,
        "nutrients": standards_nutrients,
        "ratios": [
            {"key": "ca_p", "name": "鈣磷比 (Ca:P)", "numerator": "ca_mg", "denominator": "p_mg", "ideal_min": 1.0, "ideal_max": 2.0},
            {"key": "omega_6_3", "name": "Omega-6 : Omega-3", "numerator": "omega6_g", "denominator": "omega3_g", "ideal_min": 5.0, "ideal_max": 10.0},
            {"key": "zn_cu", "name": "鋅銅比 (Zn:Cu)", "numerator": "zn_mg", "denominator": "cu_mg", "ideal_min": 8.0, "ideal_max": 15.0},
            {"key": "ca_per_1000kcal", "name": "鈣 / 1000 kcal", "numerator": "ca_mg", "denominator": "kcal", "scale": 1000},
            {"key": "p_per_1000kcal", "name": "磷 / 1000 kcal", "numerator": "p_mg", "denominator": "kcal", "scale": 1000},
            {"key": "protein_per_1000kcal", "name": "蛋白質 / 1000 kcal", "numerator": "protein", "denominator": "kcal", "scale": 1000},
            {"key": "fat_per_1000kcal", "name": "脂肪 / 1000 kcal", "numerator": "fat", "denominator": "kcal", "scale": 1000},
            {"key": "metcys_lys", "name": "Met+Cys : Lys", "numerator": "met_cys_g", "denominator": "lys_g", "ideal_min": 0.5, "ideal_max": 1.0},
            {"key": "lys_arg", "name": "Lys : Arg", "numerator": "lys_g", "denominator": "arg_g"},
            {"key": "leu_ile", "name": "Leu : Ile", "numerator": "leu_g", "denominator": "ile_g"},
            {"key": "met_cys", "name": "Met : Cys", "numerator": "met_g", "denominator": "cys_g"},
        ]
    }
    
    with open(OUT_STANDARDS, "w", encoding="utf-8") as f:
        json.dump(standards, f, ensure_ascii=False, indent=1)
    print(f"✓ {OUT_STANDARDS}: {len(standards_nutrients)} nutrients + {len(standards['ratios'])} ratios")

if __name__ == "__main__":
    main()
